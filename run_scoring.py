"""
AutoEIT Test II: Automated EIT Scoring
Applies the Ortega (2000) meaning-based rubric to score learner transcriptions.
"""

import os
import sys
import json
import openpyxl

sys.path.insert(0, os.path.dirname(__file__))
from src.score_rules import score_eit, clean, content_overlap, levenshtein_sim

# ============================================================
# CONFIG
# ============================================================

INPUT_PATH = "data/AutoEIT_Sample_Transcriptions_for_Scoring.xlsx"
OUTPUT_DIR = "output"
OUTPUT_XLSX = os.path.join(OUTPUT_DIR, "AutoEIT_Scores_Complete.xlsx")

PARTICIPANTS = ["38001-1A", "38002-2A", "38004-2A", "38006-2A"]

TARGETS = [
    "Quiero cortarme el pelo",
    "El libro está en la mesa",
    "El carro lo tiene Pedro",
    "El se ducha cada mañana",
    "Qué dice usted que va a hacer hoy",
    "Dudo que sepa manejar muy bien",
    "Las calles de esta ciudad son muy anchas",
    "Puede que llueva mañana todo el día",
    "Las casas son muy bonitas pero caras",
    "Me gustan las películas que acaban bien",
    "El chico con el que yo salgo es español",
    "Después de cenar me fui a dormir tranquilo",
    "Quiero una casa en la que vivan mis animales",
    "A nosotros nos fascinan las fiestas grandiosas",
    "Ella sólo bebe cerveza y no come nada",
    "Me gustaría que el precio de las casas bajara",
    "Cruza a la derecha y después sigue todo recto",
    "Ella ha terminado de pintar su apartamento",
    "Me gustaría que empezara a hacer más calor pronto",
    "El niño al que se le murió el gato está triste",
    "Una amiga mía cuida a los niños de mi vecino",
    "El gato que era negro fue perseguido por el perro",
    "Antes de poder salir él tiene que limpiar su cuarto",
    "La cantidad de personas que fuman ha disminuido",
    "Después de llegar a casa del trabajo tomé la cena",
    "El ladrón al que atrapó la policía era famoso",
    "Le pedí a un amigo que me ayudara con la tarea",
    "El examen no fue tan difícil como me habían dicho",
    "Serías tan amable de darme el libro que está en la mesa",
    "Hay mucha gente que no toma nada para el desayuno",
]


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(INPUT_PATH)

    all_scores = {}

    for sheet_name in PARTICIPANTS:
        ws = wb[sheet_name]
        scores = []

        print(f"\n{'='*70}")
        print(f"PARTICIPANT {sheet_name}")
        print(f"{'='*70}")
        print(f"{'#':>2} | {'Score':>5} | {'CntOv':>5} | {'LevSim':>6} | Transcription")
        print("-" * 80)

        for i in range(30):
            row = i + 2
            transcription = ws.cell(row=row, column=3).value or ""
            target = TARGETS[i]

            score = score_eit(target, transcription)
            c_ov = content_overlap(target, transcription)
            l_sim = levenshtein_sim(target, transcription)

            scores.append({
                "item": i + 1,
                "target": target,
                "transcription": transcription,
                "score": score,
                "content_overlap": round(c_ov, 2),
                "levenshtein_sim": round(l_sim, 2),
            })

            # Write score to column D
            ws.cell(row=row, column=4, value=score)

            print(f"{i+1:>2} |   {score}   | {c_ov:.2f} | {l_sim:.4f} | {transcription[:60]}")

        all_scores[sheet_name] = scores

        # Summary
        avg = sum(s["score"] for s in scores) / 30
        dist = [sum(1 for s in scores if s["score"] == v) for v in range(5)]
        print(f"\nAverage: {avg:.2f} | Distribution: 0={dist[0]} 1={dist[1]} 2={dist[2]} 3={dist[3]} 4={dist[4]}")

    # Save scored Excel
    wb.save(OUTPUT_XLSX)
    print(f"\nScores saved to {OUTPUT_XLSX}")

    # Save JSON details
    with open(os.path.join(OUTPUT_DIR, "scores_detailed.json"), "w", encoding="utf-8") as f:
        json.dump(all_scores, f, ensure_ascii=False, indent=2)
    print("Detailed results saved to output/scores_detailed.json")


if __name__ == "__main__":
    main()